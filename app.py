
import streamlit as st

##Titulo global del app
st.image('./logo.png', width= 500)


#Menu del streamlit
menu = ['Inicio', 'Informes UA']

choice = st.sidebar.selectbox('Menu', menu)


if choice == 'Inicio':
    st.markdown('### Conjunto de scripts automatizados para TP informatica')

elif choice == 'Informes UA':
    
    import io
    import os
    import pandas as pd
    import numpy as np
    import shutil
    import subprocess
    from openpyxl import load_workbook
    
    listado = st.file_uploader("Listado de facturas (xlsx)", type=["xlsx"])
    plantilla = st.file_uploader("Plantilla (xlsx)", type=["xlsx"])
    
    if (listado is not None) and (plantilla is not None):

        #Cargamos el listado del gestor de facturas
        listado = pd.read_excel(listado, skiprows=9)

        #Limpiar campos vacios
        listado.dropna(axis = 1, how='all', inplace = True)
        clean = listado.iloc[:, 1:].copy()
        clean['Unnamed: 7'] = clean['Unnamed: 7'].replace('.', np.nan)
        clean['Unnamed: 7'] = clean['Unnamed: 7'].replace('Artículo', np.nan)
        clean.dropna(subset = ['Unnamed: 7'], inplace = True)
        clean['Unnamed: 2'] = clean['Unnamed: 2'].astype(str).str[:-2]

        #Cambio column names
        cols = ['Nº CLIENTE', 'NOMBRE CLIENTE', 'REFERENCIA ARTÍCULO', 
            'DESCRIPCIÓN ARTÍCULO', 'CANTIDAD SERVIDA', 'PRECIO UNITARIO',
            'IMPORTE', 'Nº FACTURA', 'FECHA']

        clean.columns = cols
        clean.reset_index(drop=True, inplace = True)


        cols_num = ['CANTIDAD SERVIDA', 'PRECIO UNITARIO', 'IMPORTE']
        for col in cols_num:
            clean[col] = clean[col].str.replace('.', '')
            clean[col] = clean[col].str.replace(',', '.').astype(float)

        #Calculo nuevas columnas necesarias para el Report UA
        new_cols = ['UNIDAD DE VENTA', 'IMPORTE IVA', 'TOTAL VENTA', 'FECHA DE PEDIDO']
        for col in new_cols:
            clean[col] =pd.Series([], dtype=float)


        cols_ord = ['Nº CLIENTE', 'NOMBRE CLIENTE', 'REFERENCIA ARTÍCULO', 
            'DESCRIPCIÓN ARTÍCULO', 'UNIDAD DE VENTA', 'CANTIDAD SERVIDA', 'PRECIO UNITARIO',
            'IMPORTE', 'IMPORTE IVA', 'TOTAL VENTA','FECHA DE PEDIDO', 'Nº FACTURA', 'FECHA']
        df = clean[cols_ord].copy()


        df['UNIDAD DE VENTA'] = df['CANTIDAD SERVIDA'].copy()
        df['FECHA DE PEDIDO'] = df['FECHA'].copy()
        df['IMPORTE IVA'] = df['IMPORTE'] * 0.21
        df['IMPORTE IVA'] = df['IMPORTE IVA'].round(3)
        df['TOTAL VENTA'] = df['IMPORTE'] + df['IMPORTE IVA']

        #Insertar datos en el excel
        #Creamos copia del archivo plantilla

        global excel_file
        excel_file = 'ReportCompleted.xlsx'

        # Cargar el archivo de plantilla y la hoja específica
        book = load_workbook(plantilla)
        
        #Seleccion plantilla y hoja
        sheet_name = book.sheetnames[0]

        # Seleccionar la hoja donde se va a escribir el DataFrame
        sheet = book[sheet_name]

        # Encontrar la primera columna vacía en la fila 10
        start_col = 1
        while sheet.cell(row=7, column=start_col).value is not None:
            start_col += 1

        # Escribir el DataFrame en el archivo de Excel a partir de la fila 10 y la columna encontrada, sin incluir los nombres de las columnas
        for index, row in df.iterrows():
            for col_num, value in enumerate(row, start=start_col):
                sheet.cell(row=index + 7, column=col_num, value=value)

        # Guardar y cerrar el archivo de Excel
        book.save(excel_file)
        book.close()

        with open("ReportCompleted 131-136 1 Semestre.xlsx", "rb") as file:

            excel_data = io.BytesIO(file.read())

            btn = st.download_button(
                    label="Descargar",
                    data=excel_data.getvalue(),
                    file_name = 'ReportCompleted.xlsx', 
                    mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                  )
            
    else:
        st.button('Descargar')
        

